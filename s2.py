import cv2
import os
import datetime
import openpyxl

# Get the person ID
person_id = input("Enter person ID: ")

# Create a folder to save the extracted faces
if not os.path.exists('faces'):
    os.makedirs('faces')

# Create a folder for the person if it doesn't exist
person_folder = f'faces/{person_id}'
if not os.path.exists(person_folder):
    os.makedirs(person_folder)

# Load the saved face images
face_images = []
for filename in os.listdir(person_folder):
    img = cv2.imread(os.path.join(person_folder, filename))
    face_images.append(img)

# Create a VideoCapture object to read the camera
cap = cv2.VideoCapture(0)

attendance = {}

while True:
    ret, frame = cap.read()
    if not ret:
        print("Can't receive frame (stream end?). Exiting...")
        break

    # Convert the frame to grayscale
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Use OpenCV's Haar cascade classifier to detect faces
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    # Iterate over the detected faces
    for (x, y, w, h) in faces:
        # Extract the face from the frame
        face = frame[y:y + h, x:x + w]

        # Compare the extracted face with the saved face images
        for img in face_images:
            # Resize the face and image to the same size
            img = cv2.resize(img, (face.shape[1], face.shape[0]))

            # Convert the face and image to grayscale
            face_gray = cv2.cvtColor(face, cv2.COLOR_BGR2GRAY)
            img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            # Calculate the absolute difference between the face and image
            diff = cv2.absdiff(face_gray, img_gray)

            # Calculate the similarity score (lower score means higher similarity)
            score = cv2.norm(diff)

            # Calculate the percentage similarity
            percentage = (1 - score / (face_gray.shape[0] * face_gray.shape[1])) * 100

            # Check if the percentage similarity is above a certain threshold
            if percentage > 80:
                # Mark attendance
                current_time = datetime.datetime.now()
                attendance[current_time.strftime("%Y-%m-%d")] = current_time.strftime("%H:%M:%S")
                print(
                    f'Attendance marked for {person_id} at {current_time.strftime("%H:%M:%S")} on {current_time.strftime("%Y-%m-%d")}')
            cv2.imshow('Video Frame', frame)

    # Wait for a key press
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()

print("Attendance:")
for date, time in attendance.items():
    print(f"{date}: {time}")

# Create an Excel workbook and sheet
wb = openpyxl.Workbook()
sheet = wb.active

# Set the header row
sheet['A1'] = 'Name'
sheet['B1'] = 'Date'
sheet['C1'] = 'Time'
sheet['D1'] = 'Status'

row = 2
for date, time in attendance.items():
    # Write attendance data to the sheet
    sheet[f'A{row}'] = person_id
    sheet[f'B{row}'] = date
    sheet[f'C{row}'] = time
    cell = sheet[f'D{row}']
    cell.value = 'Present'
    cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', fill_type='solid')
    row += 1

# Mark today's date as present
today = datetime.date.today().strftime("%Y-%m-%d")
time = datetime.datetime.now().strftime("%H:%M:%S")
if today not in attendance:
    sheet[f'A{row}'] = person_id
    sheet[f'B{row}'] = today
    sheet[f'C{row}'] = time
    cell = sheet[f'D{row}']
    cell.value = 'Present'
    cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', fill_type='solid')
    row += 1

# Save the Excel file
wb.save(f'attendance_student.xlsx')

print("Attendance written to Excel file!")